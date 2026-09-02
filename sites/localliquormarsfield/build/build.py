#!/usr/bin/env python3
"""Regenerate spirits.html and range.html from the product export.

    python3 build/build.py                     # no prices (current live state)
    python3 build/build.py --prices            # step 2: include prices
    python3 build/build.py --source other.xlsx # use a fresher export

Curation rule that matters: price is only a quality signal where pack size is
constant. Whisky/gin/tequila sort well on price; beer and premixed do not,
because a high price there means a carton, not a better product. So beer is
curated on style and awards, and premixed is excluded entirely.
"""
import argparse, collections, html, json, pathlib, re, sys

HERE = pathlib.Path(__file__).resolve().parent
SITE = HERE.parent
DEFAULT_SOURCE = SITE.parent.parent / "seo/trafalgar/range-data/liquor-products.xlsx"

e = html.escape
AC = ' aria-current="page"'
NAV = [("/", "Home", "home"), ("/spirits.html", "Spirits", "spirits"),
       ("/range.html", "Full range", "range"), ("/visit.html", "Visit us", "visit"),
       ("/about.html", "About", "about"), ("/contact.html", "Contact", "contact")]

PRICE_SORTS_QUALITY = {"Whisky", "Gin", "Tequila", "Rum", "Cognac", "Brandy", "Vodka",
                       "Liqueurs", "Other Spirits", "Red Wine", "White Wine",
                       "Champagne & Sparkling", "Fortified Wine"}
SPIRIT_SUBS = ("Whisky", "Gin", "Tequila", "Rum", "Cognac", "Brandy", "Vodka", "Liqueurs")


def load(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True, data_only=True)["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    hdr = [h.strip() if isinstance(h, str) else h for h in next(rows)]
    seen, cols = collections.Counter(), []
    for h in hdr:                       # 'Tags' appears twice in the export
        seen[h] += 1
        cols.append(h if seen[h] == 1 else f"{h}_{seen[h]}")
    return [dict(zip(cols, r)) for r in rows if any(r)]


def s(d, f):
    v = d.get(f)
    return "" if v is None else str(v).strip()


def price(d):
    try:
        return float(d.get("Price"))
    except (TypeError, ValueError):
        return 0.0


def is_premium(d):
    cat, sub, p = s(d, "Product category"), s(d, "Sub category"), price(d)
    if sub == "Premixed Drinks":
        return False
    if cat == "Spirits":
        return p >= 60 or bool(s(d, "Age")) or bool(s(d, "Awards Won"))
    if cat == "Wine":
        return p >= 50 or (s(d, "Awards Won") and p >= 30)
    if cat == "Beer & Cider":
        return sub == "Craft Beer" and (bool(s(d, "Awards Won")) or p >= 70)
    return False


def tidy(name):
    n = re.sub(r"\s+", " ", name).strip()
    return re.sub(r"\s*\d+(\.\d+)?\s*(mL|ML|L)\b.*$", "", n).strip()


def curate(data):
    sel = [d for d in data if is_premium(d)]
    out = {}
    for sub in list(PRICE_SORTS_QUALITY) + ["Craft Beer"]:
        items = sorted((d for d in sel if s(d, "Sub category") == sub), key=lambda d: -price(d))
        if not items:
            continue
        picks, seen_brand = [], set()
        for d in items:                 # one per brand, prestige first
            b = s(d, "Brand")
            if not b or b in seen_brand:
                continue
            seen_brand.add(b)
            picks.append({"name": tidy(s(d, "Product name")), "brand": b,
                          "origin": s(d, "Region Of Origin") or s(d, "Country Of Origin"),
                          "price": price(d), "award": bool(s(d, "Awards Won"))})
            if len(picks) >= 12:
                break
        out[sub] = {"count": len(items), "picks": picks,
                    "brands": [b for b, _ in collections.Counter(
                        s(d, "Brand") for d in items if s(d, "Brand")).most_common(10)],
                    "awarded": sum(1 for d in items if s(d, "Awards Won"))}
    return sel, out


def nav(cur):
    return "".join('\n        <li><a href="%s"%s>%s</a></li>' % (h, AC if k == cur else "", t)
                   for h, t, k in NAV)


def head(title, desc, path, cur):
    return ('<!doctype html>\n<html lang="en-AU">\n<head>\n<meta charset="utf-8">\n'
            '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
            f'<title>{title}</title>\n<meta name="description" content="{desc}">\n'
            f'<link rel="canonical" href="https://localliquormarsfield.com.au{path}">\n'
            '<meta property="og:type" content="website">\n'
            '<meta property="og:site_name" content="Local Liquor Marsfield">\n'
            f'<meta property="og:title" content="{title}">\n'
            f'<meta property="og:description" content="{desc}">\n'
            f'<meta property="og:url" content="https://localliquormarsfield.com.au{path}">\n'
            '<meta property="og:locale" content="en_AU">\n'
            '<link rel="stylesheet" href="/assets/style.css">\n</head>\n<body>\n'
            '<header class="masthead">\n  <div class="wrap">\n'
            '    <a class="brand" href="/">Local Liquor Marsfield'
            '<span>Bottle shop &middot; Marsfield NSW</span></a>\n'
            f'    <nav aria-label="Main">\n      <ul>{nav(cur)}\n      </ul>\n'
            '    </nav>\n  </div>\n</header>\n<main class="wrap">\n')


FOOT = (SITE / "build" / "footer.html")


def section(F, key, heading, blurb, limit=10, prices=False):
    d = F.get(key)
    if not d:
        return ""
    lis = []
    for x in d["picks"][:limit]:
        origin = ' <em>(%s)</em>' % e(x["origin"]) if x["origin"] else ""
        pr = ' &mdash; <strong>$%s</strong>' % (f"{x['price']:.0f}" if x["price"] >= 1 else "") \
             if prices and x["price"] >= 1 else ""
        lis.append("\n    <li><strong>%s</strong> &mdash; %s%s%s</li>"
                   % (e(x["brand"]), e(x["name"]), origin, pr))
    aw = " %d of them have won a medal." % d["awarded"] if d["awarded"] else ""
    anchor = key.lower().replace(" ", "-").replace("&", "and")
    return ('\n<h2 id="%s">%s</h2>\n<p>%s%s</p>\n'
            '<p><strong>Brands we carry include:</strong> %s.</p>\n<ul>%s\n</ul>\n'
            % (anchor, heading, blurb, aw, ", ".join(e(b) for b in d["brands"][:8]), "".join(lis)))


def price_note(prices):
    if prices:
        return ('\n<div class="note">\n  <p>\n    Prices shown were current when this page '
                'was last updated and can change. Ring\n    <a href="tel:+61298681070">'
                '(02) 9868 1070</a> to confirm before making a trip.\n  </p>\n</div>\n')
    return ('\n<div class="note">\n  <p>\n    Prices are not listed here because they move. Ring\n'
            '    <a href="tel:+61298681070">(02) 9868 1070</a> and we will give you today&rsquo;s\n'
            '    price and check the bottle is on the shelf while you wait.\n  </p>\n</div>\n')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", default=str(DEFAULT_SOURCE))
    ap.add_argument("--prices", action="store_true", help="include prices (step 2)")
    a = ap.parse_args()

    data = load(a.source)
    sel, F = curate(data)
    foot = FOOT.read_text()
    sp = sum(F[k]["count"] for k in SPIRIT_SUBS if k in F)

    # ---- spirits.html ----
    o = head("Spirits | Local Liquor Marsfield &mdash; whisky, gin &amp; tequila",
             "Over %d premium spirits at Local Liquor Marsfield, Shop 5A Trafalgar Place, "
             "Marsfield NSW. Single malt scotch, Japanese whisky, Australian gin, agave tequila." % sp,
             "/spirits.html", "spirits")
    o += ('\n<h1>Spirits</h1>\n<p class="lede">\n  Spirits are what this shop is built around. '
          'Over %d premium bottles on the\n  shelf &mdash; single malt scotch, Japanese whisky, '
          'Australian botanical gin and\n  100%% agave tequila &mdash; alongside the everyday '
          'pouring bottles.\n</p>\n\n<div class="actions">\n'
          '  <a class="btn btn-primary" href="tel:+61298681070">Ask what&rsquo;s in stock</a>\n'
          '  <a class="btn btn-secondary" href="https://g.page/r/CenxQG-m-gUkEBI">Get directions</a>\n'
          '</div>\n' % sp)
    o += price_note(a.prices)
    for k, h2, bl, lim in [
        ("Whisky", "Whisky", "Our deepest shelf. Speyside, Highland and Islay single malts, aged blends, plus Japanese, Irish and American whiskey.", 10),
        ("Gin", "Gin", "London dry through to Australian botanical gins, with a growing local shelf.", 10),
        ("Tequila", "Tequila &amp; agave", "Blanco, reposado and a&ntilde;ejo, including 100% agave expressions.", 10),
        ("Rum", "Rum", "Aged, spiced and dark rums from the Caribbean and Australia.", 10),
        ("Vodka", "Vodka", "Everyday through to small-batch and imported.", 10),
        ("Liqueurs", "Liqueurs", "Amaro, cream, coffee and fruit liqueurs for cocktails and after dinner.", 10),
        ("Cognac", "Cognac &amp; brandy", "VS through XO, plus Australian brandy.", 8)]:
        o += section(F, k, h2, bl, lim, a.prices)
    o += ('\n<h2>Looking for something specific?</h2>\n<p>\n  We carry a great deal more than fits '
          'on this page, and we can usually order in\n  what we do not stock. Ring '
          '<a href="tel:+61298681070">(02) 9868 1070</a> &mdash;\n  someone will walk the aisle '
          'for you.\n</p>\n')
    (SITE / "spirits.html").write_text(o + foot)

    # ---- range.html ----
    o = head("Our range | Local Liquor Marsfield &mdash; spirits, craft beer &amp; wine",
             "Premium spirits, 450+ craft beers and a curated wine selection at Local Liquor "
             "Marsfield, Shop 5A Trafalgar Place, Marsfield NSW 2122.", "/range.html", "range")
    o += ('\n<h1>Our range</h1>\n<p class="lede">\n  Around %s lines in store. Below is the pick of '
          'it &mdash; the premium\n  spirits, the craft beer and the wines worth making room for. '
          'The everyday\n  bottles are all on the shelf too; this page is the good stuff.\n</p>\n\n'
          '<div class="actions">\n  <a class="btn btn-primary" href="/spirits.html">Browse spirits</a>\n'
          '  <a class="btn btn-secondary" href="tel:+61298681070">Ring the shop</a>\n</div>\n\n'
          '<h2>Spirits &mdash; where we go deepest</h2>\n<p>\n  Over %d premium spirits, led by %d '
          'whiskies and %d gins. Single malt scotch by\n  region, Japanese whisky, Australian '
          'botanical gin and agave tequila.\n</p>\n'
          '<p><a href="/spirits.html"><strong>See the full spirits range &rarr;</strong></a></p>\n\n'
          '<h2>Craft beer</h2>\n<p>\n  Over %d craft beers &mdash; a deeper independent range than '
          'most bottle shops\n  our size carry, and the first thing we would point a chain shopper '
          'to.\n  %d of the ones we feature have won a medal.\n</p>\n'
          '<p><strong>Breweries include:</strong> %s.</p>\n'
          '<p>\n  Alongside them: around 270 Australian mainstream lines, 120 imports, 130 ciders,\n'
          '  and roughly 50 low-carb, mid-strength and zero-alcohol beers for whoever is driving.\n</p>\n\n'
          '<h2>Wine &mdash; a tighter selection</h2>\n<p>\n  There are over 2,100 wines in store. '
          'Rather than list them all, these are the\n  ones we would actually recommend: %d reds, '
          '%d whites and %d sparkling that are\n  either award winners or genuinely worth the '
          'money.\n</p>\n'
          % (f"{round(len(data), -2):,}", sp, F["Whisky"]["count"], F["Gin"]["count"],
             450, F["Craft Beer"]["awarded"],
             ", ".join(e(b) for b in F["Craft Beer"]["brands"][:10]),
             F["Red Wine"]["count"], F["White Wine"]["count"], F["Champagne & Sparkling"]["count"]))
    for k, h2, bl in [
        ("Red Wine", "Red wine", "Shiraz, cabernet, pinot and grenache &mdash; Barossa, McLaren Vale, Coonawarra and the Yarra."),
        ("White Wine", "White wine", "Chardonnay, riesling, sauvignon blanc and pinot gris from Adelaide Hills, Clare Valley and Marlborough."),
        ("Champagne & Sparkling", "Champagne &amp; sparkling", "Prosecco and Australian sparkling through to Champagne proper.")]:
        o += section(F, k, h2, bl, 8, a.prices)
    o += ('\n<h2>Also in store</h2>\n<p>\n  Fortified wine and cask, premixed cans and bottles, ice, '
          'mixers and glassware.\n  Everything you would expect from a full bottle shop, whether or '
          'not it made this page.\n</p>\n\n<h2>Can&rsquo;t see what you want?</h2>\n<p>\n  Ring '
          '<a href="tel:+61298681070">(02) 9868 1070</a>. We will check the shelf while\n  you are '
          'on the phone, and order it in if we do not have it.\n</p>\n')
    (SITE / "range.html").write_text(o + foot)

    print("featured %d of %d" % (len(sel), len(data)))
    print("prices: %s" % ("INCLUDED" if a.prices else "omitted"))
    print("wrote spirits.html, range.html")


if __name__ == "__main__":
    main()
